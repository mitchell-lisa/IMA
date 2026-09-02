"""Builds docs/marketready-question-bank.xlsx from docs/question-bank.json.

The workbook is a maintenance surface for the Copilot Excel agent: wording
columns can be edited and proposed back; ids and structure must not change.
No formulas are used, so no recalculation step is needed.
"""
import json, sys
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

src = sys.argv[1] if len(sys.argv) > 1 else "docs/question-bank.json"
dst = sys.argv[2] if len(sys.argv) > 2 else "docs/marketready-question-bank.xlsx"
data = json.load(open(src))

FONT = "Arial"
HEAD_FILL = PatternFill("solid", fgColor="0F3D5E")
EDIT_FILL = PatternFill("solid", fgColor="FFFFCC")
wb = Workbook()

def sheet(title, headers, rows, editable=(), widths=None):
    ws = wb.create_sheet(title)
    ws.append(headers)
    for c in ws[1]:
        c.font = Font(name=FONT, bold=True, color="FFFFFF"); c.fill = HEAD_FILL
        c.alignment = Alignment(vertical="top", wrap_text=True)
    for r in rows:
        ws.append(r)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.font = Font(name=FONT); c.alignment = Alignment(vertical="top", wrap_text=True)
            if headers[c.column - 1] in editable: c.fill = EDIT_FILL
    for i, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(i)].width = (widths or {}).get(h, 18)
    ws.freeze_panes = "A2"
    return ws

legend = wb.active; legend.title = "Read me"
for line in [
    "MarketReady Risk Diagnostic — question bank workbook",
    f"Generated from code on {data['generatedAt']}. Regenerate with: npm run docs:bank",
    "",
    "Yellow cells are safe to edit (wording). Ids, categories, weights, and structure are data the app depends on: propose changes, do not rename ids.",
    "Maturity scale on every question: 0 undocumented/reactive · 1 partial/inconsistent · 2 documented · 3 documented, monitored, reviewed. 'Unknown' reduces confidence, never the score.",
    "Findings are deterministic rules; the 'when' logic lives in code and is summarized in docs/scoring-spec.md.",
    "Source of truth: src/lib/diagnostic/*.ts. This workbook is an export, not the master.",
]:
    legend.append([line])
legend.column_dimensions["A"].width = 120
for row in legend.iter_rows():
    for c in row: c.font = Font(name=FONT, bold=(c.row == 1), size=12 if c.row == 1 else 10); c.alignment = Alignment(wrap_text=True)

sheet("Categories", ["id", "label", "short", "description"], [[c["id"], c["label"], c["short"], c["description"]] for c in data["categories"]], editable=("label", "short", "description"), widths={"description": 60})

ind_heads = ["id", "label", "shortLabel", "description", "naics", "employeeThreshold", "marketNote"] + [f"weight_{c['id']}" for c in data["categories"]]
sheet("Industries", ind_heads, [[i["id"], i["label"], i["shortLabel"], i["description"], ", ".join(i["naics"]), i["employeeThreshold"], i["marketNote"]] + [i["categoryWeights"][c["id"]] for c in data["categories"]] for i in data["industries"]], editable=("label", "shortLabel", "description", "marketNote"), widths={"description": 40, "marketNote": 60})

q_heads = ["order", "id", "category", "topic", "branch", "prompt", "help", "weight_cre_owner", "weight_multifamily", "weight_other", "critical_atOrBelow", "critical_message"]
sheet("Questions", q_heads, [[n + 1, q["id"], q["category"], q["topic"], q.get("branch", ""), q["prompt"], q.get("help", ""), q["weights"]["cre_owner"], q["weights"]["multifamily"], q["weights"]["other"], (q.get("critical") or {}).get("atOrBelow", ""), (q.get("critical") or {}).get("message", "")] for n, q in enumerate(data["questions"])], editable=("topic", "prompt", "help", "critical_message"), widths={"prompt": 60, "help": 50, "critical_message": 50})

sheet("Options", ["question_id", "value", "label"], [[q["id"], o["value"], o["label"]] for q in data["questions"] for o in q["options"]], editable=("label",), widths={"label": 90})

sheet("Variants", ["question_id", "industry", "prompt", "help"], [[q["id"], ind, v.get("prompt", ""), v.get("help", "")] for q in data["questions"] for ind, v in (q.get("variants") or {}).items()], editable=("prompt", "help"), widths={"prompt": 70, "help": 60})

sheet("Findings", ["id", "category", "priority", "title", "body", "detail", "question_ids"], [[f["id"], f["category"], f["priority"], f["title"], f["body"], f.get("detail", ""), ", ".join(f["questionIds"])] for f in data["findings"]], editable=("title", "body", "detail"), widths={"title": 50, "body": 60, "detail": 60, "question_ids": 40})

sheet("Modules", ["id", "name", "headline", "subhead", "audience", "focusCategories", "focusTitle", "focusIntro"], [[m["id"], m["name"], m["headline"], m["subhead"], m["audience"], ", ".join(m["focusCategories"]), m["focusTitle"], m["focusIntro"]] for m in data["modules"]], editable=("name", "headline", "subhead", "audience", "focusTitle", "focusIntro"), widths={"headline": 50, "subhead": 60, "audience": 50, "focusIntro": 60})

sheet("Niches", ["id", "label", "industry", "whyAttractive", "dynamicModule", "hasModule"], [[n["id"], n["label"], n["industry"], n["whyAttractive"], n["dynamicModule"], "yes" if n["hasModule"] else "no"] for n in data["niches"]], editable=("label",), widths={"whyAttractive": 40, "dynamicModule": 50})

tests = [
    ["all_documented_reviewed", "cre_owner", "every core answer 3", "overall 100, band strong, confidence 100"],
    ["all_undocumented", "cre_owner", "every core answer 0", "overall 0, band priority, critical flags present"],
    ["unknowns_reduce_confidence", "cre_owner", "all 3 except two answers 'unknown'", "overall 100, confidence 89"],
    ["category_without_data", "cre_owner", "claims answers all 'unknown'", "claims score null; overall from other five"],
    ["late_renewal_with_strengths", "cre_owner", "all 3 except renewal lead time 0", "finding renewal_starts_late; strengths present"],
    ["business_change_gate", "cre_owner", "business changes 0; profile recent acquisition true/false", "finding business_outpaced_governance only when true"],
    ["management_agreement_branch", "cre_owner", "usesThirdPartyManager true; management agreement 0", "branch applicable; critical flag br_management_agreement"],
    ["contracts_module_focus", "multifamily", "module=contracts; COI 0, requirements 1, vendor transfer 1", "focus card lists those three practices"],
]
sheet("Test cases", ["id", "industry", "input", "expected"], tests, widths={"input": 50, "expected": 60})

wb.save(dst)
print(f"wrote {dst}")
